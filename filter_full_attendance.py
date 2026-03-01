import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

CSV_FILE = "Student_Attendance_Table.csv"
OUTPUT_FILE = "Full_Attendance_Students.xlsx"
THRESHOLD_MINUTES = 180  # 3 hours

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
bold_font = Font(bold=True, size=11)
normal_font = Font(size=11)
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Read CSV
with open(CSV_FILE, newline="") as f:
    reader = csv.reader(f)
    headers = next(reader)
    rows = list(reader)

# Filter: students who attended ALL sessions (no '--' in any session column)
session_cols = headers[1:]  # Session 36..43
full_attendance = []
for row in rows:
    sessions = row[1:]
    if all(val.strip() != "--" for val in sessions):
        full_attendance.append(row)

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Full Attendance Students"

# Title row
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers) + 1)
title_cell = ws.cell(row=1, column=1, value="Students Who Attended All Sessions (36-43)")
title_cell.font = Font(bold=True, size=14, color="1F4E79")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Subtitle
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers) + 1)
subtitle = ws.cell(
    row=2, column=1,
    value=f"Yellow = Attended less than 3 hours (< {THRESHOLD_MINUTES} min)  |  Values are in minutes"
)
subtitle.font = Font(italic=True, size=10, color="996600")
subtitle.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 22

# Header row (row 3): Sr.No + original headers
header_row = 3
sr_cell = ws.cell(row=header_row, column=1, value="Sr. No.")
sr_cell.fill = header_fill
sr_cell.font = header_font
sr_cell.alignment = center_align
sr_cell.border = thin_border

for col_idx, h in enumerate(headers, start=2):
    cell = ws.cell(row=header_row, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

# Data rows
for row_idx, row_data in enumerate(full_attendance, start=1):
    excel_row = header_row + row_idx

    # Sr. No.
    sr_cell = ws.cell(row=excel_row, column=1, value=row_idx)
    sr_cell.font = normal_font
    sr_cell.alignment = center_align
    sr_cell.border = thin_border

    # Student Name
    name_cell = ws.cell(row=excel_row, column=2, value=row_data[0])
    name_cell.font = bold_font
    name_cell.alignment = left_align
    name_cell.border = thin_border

    # Session values
    for col_idx, val in enumerate(row_data[1:], start=3):
        minutes = int(val.strip())
        cell = ws.cell(row=excel_row, column=col_idx, value=minutes)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
        if minutes < THRESHOLD_MINUTES:
            cell.fill = yellow_fill

# Auto-fit column widths
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 38
for col_letter_idx in range(3, len(headers) + 2):
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(col_letter_idx)].width = 13

# Freeze header
ws.freeze_panes = "A4"

# Summary at bottom
summary_row = header_row + len(full_attendance) + 2
ws.cell(row=summary_row, column=1, value="Summary:").font = Font(bold=True, size=11)
ws.cell(row=summary_row + 1, column=1, value=f"Total students with full attendance: {len(full_attendance)}").font = Font(size=11)
ws.cell(row=summary_row + 2, column=1, value=f"Total sessions: {len(session_cols)}").font = Font(size=11)

wb.save(OUTPUT_FILE)
print(f"Created '{OUTPUT_FILE}' with {len(full_attendance)} students who attended all {len(session_cols)} sessions.")
print(f"Entries with < {THRESHOLD_MINUTES} minutes are highlighted in yellow.")
