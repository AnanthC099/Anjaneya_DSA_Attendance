import csv
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Session directories and their CSV files (in order)
SESSION_FILES = {
    "Session_36": "participants_87901948469_2026_01_31.csv",
    "Session_37": "participants_89208110803_2026_02_01.csv",
    "Session_38": "participants_89837366134_2026_02_07.csv",
    "Session_39": "participants_88492557764_2026_02_08.csv",
    "Session_40": "participants_84125620887_2026_02_14.csv",
    "Session_41": "participants_84598698913_2026_02_15.csv",
    "Session_42": "participants_83599210157_2026_02_22.csv",
    "Session_43": "participants_87190343408_2026_02_28.csv",
}

SESSION_LABELS = [
    "Session 36\n(31 Jan)",
    "Session 37\n(01 Feb)",
    "Session 38\n(07 Feb)",
    "Session 39\n(08 Feb)",
    "Session 40\n(14 Feb)",
    "Session 41\n(15 Feb)",
    "Session 42\n(22 Feb)",
    "Session 43\n(28 Feb)",
]

TOTAL_SESSIONS = len(SESSION_FILES)

# ── Load master student list ──
def load_students():
    students = {}
    path = os.path.join(BASE_DIR, "Total_Student_Data", "CPA_DSA_24_04_10_2025.csv")
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            email = row["Email"].strip().lower()
            first_name = row["First Name"].strip()
            last_name = row["Last Name"].strip()
            # Extract student ID from first name (e.g., "DSA24001_Himanshu")
            student_id = first_name.split("_")[0] if "_" in first_name else first_name
            display_name = first_name.split("_", 1)[1] if "_" in first_name else first_name
            full_name = f"{display_name} {last_name}"
            students[email] = {
                "id": first_name.split("_")[0] if "_" in first_name else "",
                "full_id": first_name,
                "name": full_name,
                "last_name": last_name,
                "email": email,
            }
    return students


# ── Load session attendance data ──
def load_session_attendance():
    """Returns dict: email -> {session_index: duration_minutes}"""
    attendance = {}
    for idx, (session_dir, csv_file) in enumerate(SESSION_FILES.items()):
        path = os.path.join(BASE_DIR, session_dir, csv_file)
        with open(path, newline="", encoding="utf-8") as f:
            lines = f.readlines()

        # Find the line with participant data header
        data_start = None
        for i, line in enumerate(lines):
            if line.startswith("Name (original name)"):
                data_start = i
                break

        if data_start is None:
            continue

        reader = csv.DictReader(lines[data_start:])
        for row in reader:
            email = row["Email"].strip().lower()
            duration = int(row["Total duration (minutes)"].strip())
            # Skip the host/instructor
            if email == "corecodeprogrammingacademy@gmail.com":
                continue
            # Skip non-student accounts (like Ananth Chandrasekharan)
            if email == "ananthchandrasekharan@gmail.com":
                continue
            if email not in attendance:
                attendance[email] = {}
            # If student appears multiple times, take the max duration
            if idx in attendance[email]:
                attendance[email][idx] = max(attendance[email][idx], duration)
            else:
                attendance[email][idx] = duration

    return attendance


# ── Styling constants ──
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=11)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_workbook(title, subtitle, students_data, attendance, session_count_filter=None):
    """
    Create an Excel workbook with attendance data.
    students_data: list of student dicts (sorted)
    attendance: full attendance dict
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # ── Title row ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + TOTAL_SESSIONS)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = CENTER_ALIGN

    # ── Subtitle row ──
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3 + TOTAL_SESSIONS)
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font = Font(bold=True, size=11, color="555555")
    sub_cell.alignment = CENTER_ALIGN

    # ── Header row ──
    header_row = 3
    headers = ["Sr. No.", "Student ID", "Student Name"] + SESSION_LABELS
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # ── Data rows ──
    for row_idx, student in enumerate(students_data, 1):
        data_row = header_row + row_idx
        email = student["email"]

        # Sr. No.
        cell = ws.cell(row=data_row, column=1, value=row_idx)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        cell.font = NORMAL_FONT

        # Student ID
        cell = ws.cell(row=data_row, column=2, value=student["full_id"])
        cell.alignment = LEFT_ALIGN
        cell.border = THIN_BORDER
        cell.font = NORMAL_FONT

        # Student Name
        cell = ws.cell(row=data_row, column=3, value=student["name"])
        cell.alignment = LEFT_ALIGN
        cell.border = THIN_BORDER
        cell.font = NORMAL_FONT

        # Session columns
        student_attendance = attendance.get(email, {})
        for sess_idx in range(TOTAL_SESSIONS):
            col = 4 + sess_idx
            if sess_idx in student_attendance:
                duration = student_attendance[sess_idx]
                cell = ws.cell(row=data_row, column=col, value=duration)
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER
                cell.font = NORMAL_FONT
                # Mark yellow if less than 180 minutes (3 hours)
                if duration < 180:
                    cell.fill = YELLOW_FILL
            else:
                cell = ws.cell(row=data_row, column=col, value="--")
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER
                cell.font = NORMAL_FONT

    # ── Column widths ──
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 25
    for i in range(TOTAL_SESSIONS):
        col_letter = chr(ord("D") + i)
        ws.column_dimensions[col_letter].width = 14

    # Set row height for header
    ws.row_dimensions[header_row].height = 40

    return wb


def main():
    students = load_students()
    attendance = load_session_attendance()

    # Calculate sessions attended for each student
    student_sessions = {}
    for email, info in students.items():
        sess_attended = sum(1 for i in range(TOTAL_SESSIONS) if i in attendance.get(email, {}))
        student_sessions[email] = sess_attended

    # Categorize students
    list_1_students = []  # All sessions attended (8/8)
    list_2_students = []  # Exactly 2 leaves (6/8)
    list_3_students = []  # Everyone else

    for email, info in students.items():
        sessions = student_sessions[email]
        if sessions == TOTAL_SESSIONS:
            list_1_students.append(info)
        elif sessions == TOTAL_SESSIONS - 2:
            list_2_students.append(info)
        else:
            list_3_students.append(info)

    # Sort each list by student ID
    for lst in [list_1_students, list_2_students, list_3_students]:
        lst.sort(key=lambda s: s["full_id"])

    # Print summary
    print(f"Total students in master list: {len(students)}")
    print(f"List 1 (All {TOTAL_SESSIONS} sessions): {len(list_1_students)} students")
    print(f"List 2 (Exactly 2 leaves / {TOTAL_SESSIONS - 2} sessions): {len(list_2_students)} students")
    print(f"List 3 (Remaining): {len(list_3_students)} students")
    print()

    # Generate List_1
    wb1 = create_workbook(
        title="List 1 - Students with 100% Attendance (All 8 Sessions)",
        subtitle="Sessions 36-43 | Yellow = attended less than 3 hours | '--' = absent | Duration in minutes",
        students_data=list_1_students,
        attendance=attendance,
    )
    path1 = os.path.join(BASE_DIR, "List_1.xlsx")
    wb1.save(path1)
    print(f"Saved: {path1}")

    # Generate List_2
    wb2 = create_workbook(
        title="List 2 - Students with Exactly 2 Leaves (6 out of 8 Sessions)",
        subtitle="Sessions 36-43 | Yellow = attended less than 3 hours | '--' = absent | Duration in minutes",
        students_data=list_2_students,
        attendance=attendance,
    )
    path2 = os.path.join(BASE_DIR, "List_2.xlsx")
    wb2.save(path2)
    print(f"Saved: {path2}")

    # Generate List_3
    wb3 = create_workbook(
        title="List 3 - Remaining Students",
        subtitle="Sessions 36-43 | Yellow = attended less than 3 hours | '--' = absent | Duration in minutes",
        students_data=list_3_students,
        attendance=attendance,
    )
    path3 = os.path.join(BASE_DIR, "List_3.xlsx")
    wb3.save(path3)
    print(f"Saved: {path3}")

    # Print details for each list
    print("\n=== LIST 1: Full Attendance (All 8 Sessions) ===")
    for s in list_1_students:
        print(f"  {s['full_id']} - {s['name']}")

    print(f"\n=== LIST 2: Exactly 2 Leaves ({TOTAL_SESSIONS - 2}/8 Sessions) ===")
    for s in list_2_students:
        email = s["email"]
        absent_sessions = [SESSION_LABELS[i].split("\n")[0] for i in range(TOTAL_SESSIONS)
                          if i not in attendance.get(email, {})]
        print(f"  {s['full_id']} - {s['name']} | Absent: {', '.join(absent_sessions)}")

    print(f"\n=== LIST 3: Remaining Students ===")
    for s in list_3_students:
        email = s["email"]
        attended = student_sessions[email]
        absent_sessions = [SESSION_LABELS[i].split("\n")[0] for i in range(TOTAL_SESSIONS)
                          if i not in attendance.get(email, {})]
        print(f"  {s['full_id']} - {s['name']} | Attended: {attended}/{TOTAL_SESSIONS} | Absent: {', '.join(absent_sessions)}")


if __name__ == "__main__":
    main()
